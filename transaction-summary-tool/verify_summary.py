# -*- coding: utf-8 -*-
"""Verify a generated Summary workbook against an independent read of the source.

Cross-checks with pandas (a different reader than the openpyxl path used to
build the file), dumps every written cell, and asserts the three blocks
reconcile with each other.

Usage:
    python verify_summary.py "transactions.xlsx" "transactions_Summary.xlsx"
"""
import argparse
import sys

import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

failures = []


def check(label, got, want, tol=0.005):
    if isinstance(want, float) or isinstance(got, float):
        ok = abs(float(got) - float(want)) <= tol
    else:
        ok = got == want
    print(f"  [{'OK ' if ok else 'FAIL'}] {label}: got {got:,} want {want:,}")
    if not ok:
        failures.append(label)


def main():
    ap = argparse.ArgumentParser(description="Verify a generated Summary workbook.")
    ap.add_argument("source", help="the original transaction export")
    ap.add_argument("summary", help="the generated *_Summary.xlsx")
    ap.add_argument("--sheet", default="All", help="raw data sheet in the source")
    ap.add_argument("--summary-sheet", default="Summary")
    ap.add_argument("--type-col", default="Type")
    ap.add_argument("--amount-col", default="Amount")
    ap.add_argument("--in-type", default="Payment")
    ap.add_argument("--out-type", default="Withdraw")
    args = ap.parse_args()

    # --- 1. Independent re-derivation with pandas ---------------------
    print("== independent pandas cross-check ==")
    df = pd.read_excel(args.source, sheet_name=args.sheet,
                       usecols=[args.type_col, args.amount_col])
    grp = df.groupby(args.type_col)[args.amount_col].agg(["count", "sum"])
    print(grp.to_string())
    print(f"  pandas total rows: {len(df):,}")
    p_count = int(grp.loc[args.in_type, "count"])
    p_sum = float(grp.loc[args.in_type, "sum"])
    w_count = int(grp.loc[args.out_type, "count"])
    w_sum = float(grp.loc[args.out_type, "sum"])

    # --- 2. Read back the written workbook ----------------------------
    ws = openpyxl.load_workbook(args.summary)[args.summary_sheet]
    print("\n== written cells ==")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5):
        if all(c.value is None for c in row):
            continue
        parts = []
        for c in row:
            if c.value is None:
                continue
            mark = "B" if c.font and c.font.bold else " "
            parts.append(f"{mark}{c.value!r}[{c.number_format}]")
        print(f"  r{row[0].row:>2}: " + " | ".join(parts))

    # Block 3 runs from row 18 to the bold total row at the bottom.
    total_row = ws.max_row
    month_rows = range(18, total_row)

    # --- 3. Block 2 against pandas ------------------------------------
    print("\n== block 2 vs pandas ==")
    check("In count", ws["B11"].value, p_count)
    check("In amount", ws["C11"].value, p_sum)
    check("Out count", ws["B12"].value, w_count)
    check("Out amount", ws["C12"].value, w_sum)
    check("Sum count", ws["B13"].value, p_count + w_count)
    check("Sum amount", ws["C13"].value, p_sum + w_sum)
    check("Diff count", ws["B14"].value, p_count - w_count)
    check("Diff amount", ws["C14"].value, p_sum - w_sum)

    # --- 4. Internal consistency --------------------------------------
    print(f"\n== internal consistency (months rows 18-{total_row - 1}) ==")
    for col, label in ((2, "In count"), (3, "In amount"),
                       (4, "Out count"), (5, "Out amount")):
        col_sum = sum(ws.cell(row=r, column=col).value for r in month_rows)
        if isinstance(col_sum, float):
            col_sum = round(col_sum, 2)
        check(f"block3 {label} sums to total row", col_sum,
              ws.cell(row=total_row, column=col).value)

    print("\n== block 1 vs block 2 ==")
    check("B3 In count", ws["B3"].value, ws["B11"].value)
    check("B4 In amount", ws["B4"].value, ws["C11"].value)
    check("B5 Out count", ws["B5"].value, ws["B12"].value)
    check("B6 Out amount", ws["B6"].value, ws["C12"].value)
    check("B7 net", round(ws["B7"].value, 2),
          round(ws["C11"].value - ws["C12"].value, 2))

    print("\n== block3 total vs block2 ==")
    check("month total In count", ws.cell(row=total_row, column=2).value, ws["B11"].value)
    check("month total In amount", ws.cell(row=total_row, column=3).value, ws["C11"].value)
    check("month total Out count", ws.cell(row=total_row, column=4).value, ws["B12"].value)
    check("month total Out amount", ws.cell(row=total_row, column=5).value, ws["C12"].value)

    print("\n" + ("ALL CHECKS PASSED" if not failures else f"FAILURES: {failures}"))
    sys.exit(1 if failures else 0)


if __name__ == "__main__":
    main()
