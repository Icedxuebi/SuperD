# -*- coding: utf-8 -*-
"""Build a Thai-format transaction Summary workbook from a raw transaction export.

Reads the source .xlsx read-only, aggregates by transaction type and calendar
month, and writes a new standalone workbook with three tables:

  1. จากไฟล์ Transaction ตั้งแต่เปิดร้านจนถึงปิดร้าน  (headline figures)
  2. สรุปธุรกรรม                                      (In / Out / total / net)
  3. สรุปรายเดือน                                     (month-by-month breakdown)

The source workbook is never opened for writing.

Usage:
    python build_summary.py "transactions.xlsx"
    python build_summary.py "transactions.xlsx" --inspect
    python build_summary.py "transactions.xlsx" -o "out.xlsx" --in-type Credit --out-type Debit

See README.md for the full option list.
"""
import argparse
import datetime
import os
import sys
from collections import defaultdict
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if hasattr(sys.stdout, "reconfigure"):          # Thai output on a cp874/cp1252 console
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

TH_MONTHS = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
             "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]

# Transaction Date arrives as a mix of real datetimes and text, so text is
# matched against these patterns in order.
DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y",
)

FMT_COUNT = '#,##0'
FMT_MONEY = '#,##0.00'
FMT_COUNT_UNIT = '#,##0" รายการ"'
FMT_MONEY_UNIT = '#,##0.00" บาท"'

BOX = Border(*(Side(style="thin", color="FFD9D9D9"),) * 4)
TOP_RULE = Border(left=Side(style="thin", color="FFD9D9D9"),
                  right=Side(style="thin", color="FFD9D9D9"),
                  top=Side(style="medium", color="FF000000"),
                  bottom=Side(style="thin", color="FFD9D9D9"))


def parse_date(value):
    """Normalise a cell that may hold a real datetime or a date string."""
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        text = value.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
    return None


def open_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"ABORT: sheet {sheet!r} not found. Sheets: {wb.sheetnames}")
    return wb, wb[sheet]


def resolve_columns(header, wanted):
    """Map logical names to 0-based column indexes by header text."""
    lookup = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    resolved = {}
    for role, name in wanted.items():
        key = name.strip().lower()
        if key not in lookup:
            available = [h for h in header if h is not None]
            sys.exit(f"ABORT: column {name!r} not found in sheet. Available: {available}")
        resolved[role] = lookup[key]
    return resolved


def inspect(path, sheet, cols):
    """Print sheet shape, headers, distinct types and date range. Writes nothing."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"file   : {path}")
    print(f"sheets : {wb.sheetnames}")
    if sheet not in wb.sheetnames:
        sys.exit(f"\nABORT: sheet {sheet!r} not found.")
    ws = wb[sheet]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print(f"sheet  : {sheet}")
    print(f"columns: {[h for h in header if h is not None]}")

    idx = resolve_columns(header, cols)
    types = defaultdict(lambda: [0, Decimal(0)])
    lo = hi = None
    unparsed = rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows += 1
        bucket = types[row[idx["type"]]]
        bucket[0] += 1
        amount = row[idx["amount"]]
        if isinstance(amount, (int, float)):
            bucket[1] += Decimal(str(amount))
        when = parse_date(row[idx["date"]])
        if when is None:
            unparsed += 1
        else:
            lo = when if lo is None or when < lo else lo
            hi = when if hi is None or when > hi else hi

    print(f"rows   : {rows:,}")
    print(f"dates  : {lo} -> {hi}   ({unparsed:,} unparseable)")
    print("types  :")
    for name in sorted(types, key=lambda k: -types[k][0]):
        count, total = types[name]
        print(f"         {str(name):<20} count={count:>10,}  sum={total:>18,}")
    print("\nPick the In/Out types with --in-type / --out-type, then re-run without --inspect.")


def aggregate(ws, idx, in_type, out_type):
    """Return (rows, {(year, month): [in_count, in_amt, out_count, out_amt]})."""
    monthly = defaultdict(lambda: [0, Decimal(0), 0, Decimal(0)])
    unknown = defaultdict(int)
    unparsed = rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows += 1
        when = parse_date(row[idx["date"]])
        if when is None:
            unparsed += 1
            continue
        raw = row[idx["amount"]]
        amount = Decimal(str(raw)) if raw is not None else Decimal(0)
        bucket = monthly[(when.year, when.month)]
        kind = row[idx["type"]]
        if kind == in_type:
            bucket[0] += 1
            bucket[1] += amount
        elif kind == out_type:
            bucket[2] += 1
            bucket[3] += amount
        else:
            unknown[kind] += 1

    # Bail out rather than emit a summary that silently drops rows.
    if unparsed:
        sys.exit(f"ABORT: {unparsed:,} rows had an unparseable date. "
                 f"Add the format to DATE_FORMATS and re-run.")
    if unknown:
        sys.exit(f"ABORT: rows with unexpected type values: {dict(unknown)}. "
                 f"Run with --inspect, then set --in-type / --out-type.")
    return rows, monthly


def style_header(cell):
    cell.fill = PatternFill("solid", fgColor="FF000000")
    cell.font = Font(bold=True, color="FFFFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX


def put(ws, row, col, value, fmt=None, bold=False, border=BOX):
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    return cell


def title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def write_summary(path, monthly, sheet_name="Summary"):
    months = sorted(monthly)
    in_count = sum(monthly[m][0] for m in months)
    in_amount = sum((monthly[m][1] for m in months), Decimal(0))
    out_count = sum(monthly[m][2] for m in months)
    out_amount = sum((monthly[m][3] for m in months), Decimal(0))
    net = in_amount - out_amount

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", (30, 27, 24, 20, 20)):
        ws.column_dimensions[col].width = width

    # --- Block 1: headline figures ------------------------------------
    title(ws, 1, "จากไฟล์ Transaction ตั้งแต่เปิดร้านจนถึงปิดร้าน")
    style_header(ws.cell(row=2, column=1, value="รายการ"))
    style_header(ws.cell(row=2, column=2, value="จำนวน"))
    block1 = [
        ("Transaction In", in_count, FMT_COUNT_UNIT),
        ("เงินเข้า", in_amount, FMT_MONEY_UNIT),
        ("Transaction Out", out_count, FMT_COUNT_UNIT),
        ("เงินโอนออก", out_amount, FMT_MONEY_UNIT),
        ("เงินคงเหลือสุทธิ", net, FMT_MONEY_UNIT),
    ]
    for offset, (label, value, fmt) in enumerate(block1):
        r = 3 + offset
        last = offset == len(block1) - 1
        put(ws, r, 1, label, bold=last)
        put(ws, r, 2, float(value), fmt=fmt, bold=last)

    # --- Block 2: transaction summary ---------------------------------
    title(ws, 9, "สรุปธุรกรรม")
    for col, text in enumerate(
            ["ประเภทธุรกรรม", "จำนวนรายการ (Transactions)", "มูลค่ารวม (บาท)"], start=1):
        style_header(ws.cell(row=10, column=col, value=text))
    block2 = [
        ("Transaction In", in_count, in_amount, False),
        ("Transaction Out", out_count, out_amount, False),
        ("รวม IN & Out", in_count + out_count, in_amount + out_amount, True),
        ("ส่วนต่าง (In – Out)", in_count - out_count, net, True),
    ]
    for offset, (label, count, amount, bold) in enumerate(block2):
        r = 11 + offset
        border = TOP_RULE if offset == 2 else BOX
        put(ws, r, 1, label, bold=bold, border=border)
        put(ws, r, 2, count, fmt=FMT_COUNT, bold=bold, border=border)
        put(ws, r, 3, float(amount), fmt=FMT_MONEY, bold=bold, border=border)

    # --- Block 3: monthly breakdown -----------------------------------
    title(ws, 16, "สรุปรายเดือน")
    for col, text in enumerate(
            ["เดือน", "In (รายการ)", "In (บาท)", "Out (รายการ)", "Out (บาท)"], start=1):
        style_header(ws.cell(row=17, column=col, value=text))
    r = 18
    for year, month in months:
        ic, ia, oc, oa = monthly[(year, month)]
        put(ws, r, 1, f"{TH_MONTHS[month - 1]} {year + 543}", bold=True)   # Buddhist era
        put(ws, r, 2, ic, fmt=FMT_COUNT)
        put(ws, r, 3, float(ia), fmt=FMT_MONEY)
        put(ws, r, 4, oc, fmt=FMT_COUNT)
        put(ws, r, 5, float(oa), fmt=FMT_MONEY)
        r += 1
    for col, value, fmt in ((1, "รวม", None), (2, in_count, FMT_COUNT),
                            (3, float(in_amount), FMT_MONEY), (4, out_count, FMT_COUNT),
                            (5, float(out_amount), FMT_MONEY)):
        put(ws, r, col, value, fmt=fmt, bold=True, border=TOP_RULE)

    wb.save(path)
    return {"in_count": in_count, "in_amount": in_amount, "out_count": out_count,
            "out_amount": out_amount, "net": net, "months": len(months),
            "total_row": r}


def main():
    ap = argparse.ArgumentParser(
        description="Build a Thai-format transaction Summary workbook.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    ap.add_argument("source", help="path to the source .xlsx transaction export")
    ap.add_argument("-o", "--output", help="output path (default: <source>_Summary.xlsx)")
    ap.add_argument("--sheet", default="All", help="worksheet holding the raw rows")
    ap.add_argument("--date-col", default="Transaction Date", help="date column header")
    ap.add_argument("--type-col", default="Type", help="transaction type column header")
    ap.add_argument("--amount-col", default="Amount", help="amount column header")
    ap.add_argument("--in-type", default="Payment", help="type value counted as money IN")
    ap.add_argument("--out-type", default="Withdraw", help="type value counted as money OUT")
    ap.add_argument("--summary-sheet", default="Summary", help="name of the sheet to create")
    ap.add_argument("--inspect", action="store_true",
                    help="print sheets, headers, types and date range; write nothing")
    args = ap.parse_args()

    if not os.path.isfile(args.source):
        sys.exit(f"ABORT: source not found: {args.source}")

    cols = {"date": args.date_col, "type": args.type_col, "amount": args.amount_col}

    if args.inspect:
        inspect(args.source, args.sheet, cols)
        return

    output = args.output or f"{os.path.splitext(args.source)[0]}_Summary.xlsx"

    wb, ws = open_sheet(args.source, args.sheet)
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = resolve_columns(header, cols)
    rows, monthly = aggregate(ws, idx, args.in_type, args.out_type)
    wb.close()

    stats = write_summary(output, monthly, args.summary_sheet)

    print(f"read    : {args.source}")
    print(f"rows    : {rows:,} over {stats['months']} months")
    print(f"in  ({args.in_type}) : {stats['in_count']:>10,} txns  {stats['in_amount']:>18,}")
    print(f"out ({args.out_type}): {stats['out_count']:>10,} txns  {stats['out_amount']:>18,}")
    print(f"net     : {stats['net']:>18,}")
    print(f"wrote   : {output}")


if __name__ == "__main__":
    main()
